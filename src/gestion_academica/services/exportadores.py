from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from gestion_academica.data.datos_demo import EstudianteCalificacion, MetadatosReporte, ResumenLogros


def exportar_reporte_excel(
    ruta: str | Path,
    meta: MetadatosReporte,
    estudiantes: Iterable[EstudianteCalificacion],
    resumen: Iterable[ResumenLogros],
) -> Path:
    """Exporta el reporte académico a un archivo XLSX."""
    estudiantes = list(estudiantes)
    resumen = list(resumen)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append(["INSTITUCIÓN", meta.institucion])
    ws.append(["DOCENTE", meta.docente])
    ws.append(["ASIGNATURA", meta.asignatura])
    ws.append(["PARALELO", meta.paralelo])
    ws.append(["TRIMESTRE", meta.trimestre])
    ws.append(["TUTOR", meta.tutor])
    ws.append([])

    encabezados = [
        "N°",
        "Nómina",
        "Aportes",
        "70%",
        "Proy. Inter.",
        "15%",
        "Examen",
        "15%",
        "Promedio final",
        "Cualitativa",
        "Observación",
    ]
    ws.append(encabezados)

    for celda in ws[8]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(fill_type="solid", fgColor="2E75B6")
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for est in estudiantes:
        ws.append(
            [
                est.numero,
                est.nombre,
                est.cal_aportes,
                est.pond_aportes,
                est.cal_proyecto,
                est.pond_proyecto,
                est.cal_examen,
                est.pond_examen,
                est.promedio_final,
                est.cualitativa,
                est.observacion,
            ]
        )

    fila_resumen_inicio = ws.max_row + 2
    ws.cell(fila_resumen_inicio, 1, "RESUMEN")
    ws.cell(fila_resumen_inicio, 1).font = Font(bold=True)
    ws.append(["Nivel", "Descripción", "Cantidad", "%"])

    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True)
        celda.fill = PatternFill(fill_type="solid", fgColor="E9DDB8")
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for item in resumen:
        ws.append([item.categoria, item.descripcion, item.cantidad, item.porcentaje])

    widths = {
        "A": 8,
        "B": 42,
        "C": 12,
        "D": 10,
        "E": 14,
        "F": 10,
        "G": 12,
        "H": 10,
        "I": 14,
        "J": 12,
        "K": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=9, max_row=fila_resumen_inicio - 2, min_col=3, max_col=9):
        for c in row:
            c.number_format = "0.00"
            c.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=9, max_row=fila_resumen_inicio - 2, min_col=1, max_col=11):
        for c in row:
            if c.column == 2:
                c.alignment = Alignment(horizontal="left")
            elif c.column > 2:
                c.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=fila_resumen_inicio + 2, max_row=ws.max_row, min_col=4, max_col=4):
        for c in row:
            c.number_format = "0%"
            c.alignment = Alignment(horizontal="center")

    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    return ruta
