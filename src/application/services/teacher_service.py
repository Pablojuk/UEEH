"""Servicio de negocio para gestión de docentes."""

from __future__ import annotations

import sqlite3
import uuid
from pathlib import Path

from src.infrastructure.persistence.repositories import DocentesRepository


class TeacherService:
    """Gestiona operaciones de docentes."""

    def __init__(self, connection: sqlite3.Connection) -> None:
        self.connection = connection
        self.repo = DocentesRepository(connection)

    def crear_docente(self, data: dict) -> None:
        payload = {
            "id_docente": data["id_docente"],
            "nombres": data["nombres"],
            "apellidos": data["apellidos"],
            "identificacion": data["identificacion"],
            "titulo": data.get("titulo") or "No registrado",
            "activo": data.get("activo", 1),
        }
        self.repo.crear(payload)

    def obtener_docente(self, id_docente: str) -> dict | None:
        return self.repo.obtener_por_id(id_docente)

    def listar_docentes(self) -> list[dict]:
        return self.repo.listar()

    def actualizar_docente(self, id_docente: str, data: dict) -> None:
        self.repo.actualizar(id_docente, data)

    def activar_docente(self, id_docente: str) -> None:
        self.repo.actualizar(id_docente, {"activo": 1})

    def inactivar_docente(self, id_docente: str) -> None:
        self.repo.actualizar(id_docente, {"activo": 0})

    def eliminar_docente(self, id_docente: str) -> tuple[bool, str]:
        asignaciones = self.connection.execute(
            "SELECT COUNT(1) AS total FROM asignaciones_docente WHERE docente_id = ?",
            (id_docente,),
        ).fetchone()
        if asignaciones and int(asignaciones["total"]) > 0:
            return False, "No se puede eliminar: el docente tiene asignaciones activas."
        self.repo.eliminar(id_docente)
        return True, "Docente eliminado correctamente."

    def importar_desde_excel(self, file_path: str) -> dict:
        path = Path(file_path)
        summary = {
            "total_leidos": 0,
            "validos": 0,
            "importados": 0,
            "omitidos": 0,
            "duplicados": 0,
            "errores": [],
        }
        rows = self._leer_archivo(path)
        summary["total_leidos"] = len(rows)
        existentes = {row.get("identificacion", "").strip().lower(): row for row in self.listar_docentes() if row.get("identificacion")}
        ids_existentes = {row.get("id_docente", "").strip().lower() for row in self.listar_docentes()}
        ids_en_archivo: set[str] = set()
        identificaciones_en_archivo: set[str] = set()

        for idx, row in enumerate(rows, start=2):
            payload = self._normalizar_docente_row(row)
            if not payload:
                summary["omitidos"] += 1
                summary["errores"].append(f"Fila {idx}: faltan columnas mínimas (id_docente, nombres, apellidos, identificacion)")
                continue
            summary["validos"] += 1
            id_norm = payload["id_docente"].lower()
            ident_norm = payload["identificacion"].lower()

            if id_norm in ids_existentes or id_norm in ids_en_archivo:
                summary["omitidos"] += 1
                summary["duplicados"] += 1
                summary["errores"].append(f"Fila {idx}: docente duplicado por ID")
                continue
            if ident_norm in existentes or ident_norm in identificaciones_en_archivo:
                summary["omitidos"] += 1
                summary["duplicados"] += 1
                summary["errores"].append(f"Fila {idx}: docente duplicado por identificación")
                continue
            try:
                self.crear_docente(payload)
                summary["importados"] += 1
                ids_en_archivo.add(id_norm)
                identificaciones_en_archivo.add(ident_norm)
            except Exception as exc:  # noqa: BLE001
                summary["omitidos"] += 1
                summary["errores"].append(f"Fila {idx}: {exc}")

        return summary

    def _leer_archivo(self, path: Path) -> list[dict]:
        if not path.exists():
            raise FileNotFoundError("Archivo no encontrado")
        if path.suffix.lower() == ".csv":
            import csv

            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                return [dict(r) for r in reader]
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            values = list(ws.iter_rows(values_only=True))
            if not values:
                return []
            headers = [str(v).strip() if v is not None else "" for v in values[0]]
            rows: list[dict] = []
            for row in values[1:]:
                item: dict[str, object] = {}
                for i, h in enumerate(headers):
                    if h:
                        item[h] = row[i] if i < len(row) else None
                if any(v not in (None, "") for v in item.values()):
                    rows.append(item)
            return rows
        finally:
            wb.close()

    @staticmethod
    def _normalizar_docente_row(raw: dict) -> dict | None:
        keys = {str(k).strip().lower(): k for k in raw.keys()}

        def get(*aliases: str) -> str:
            for alias in aliases:
                source = keys.get(alias)
                if source is not None:
                    value = raw.get(source)
                    return str(value).strip() if value is not None else ""
            return ""

        docente_id = get("id_docente", "id", "codigo_docente", "codigo")
        nombres = get("nombres", "nombre")
        apellidos = get("apellidos", "apellido")
        identificacion = get("identificacion", "identificación", "cedula", "cédula", "dni")

        if not docente_id:
            docente_id = f"DOC-{uuid.uuid4().hex[:8].upper()}"
        titulo = get("titulo", "título")

        if not (docente_id and nombres and apellidos and identificacion):
            return None

        return {
            "id_docente": docente_id,
            "nombres": nombres,
            "apellidos": apellidos,
            "identificacion": identificacion,
            "titulo": titulo or "No registrado",
            "activo": 1,
        }
