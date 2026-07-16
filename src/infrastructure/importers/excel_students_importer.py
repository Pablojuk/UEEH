"""Lectura segura de estudiantes desde Excel/CSV sin macros."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


class ExcelStudentsImporter:
    """Importador de filas crudas para estudiantes."""

    SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".csv"}

    def read_rows(self, file_path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
        path = Path(file_path)
        if not path.exists() or not path.is_file():
            raise FileNotFoundError(f"Archivo no encontrado: {file_path}")

        suffix = path.suffix.lower()
        if suffix not in self.SUPPORTED_SUFFIXES:
            raise ValueError("Formato no soportado. Use .xlsx, .xlsm o .csv")

        if suffix == ".csv":
            return self._read_csv(path)

        return self._read_excel(path, sheet_name=sheet_name)

    @staticmethod
    def _read_csv(path: Path) -> list[dict[str, Any]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise ValueError("El archivo CSV no contiene encabezados")
            return [dict(row) for row in reader]

    @staticmethod
    def _read_excel(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ImportError("openpyxl es requerido para importar Excel") from exc

        workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]

            values = list(worksheet.iter_rows(values_only=True))
            if not values:
                raise ValueError("La hoja seleccionada está vacía")

            headers = [str(value).strip() if value is not None else "" for value in values[0]]
            if not any(headers):
                raise ValueError("No se detectaron encabezados válidos")

            rows: list[dict[str, Any]] = []
            for row in values[1:]:
                record: dict[str, Any] = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    record[header] = row[index] if index < len(row) else None
                if any(value not in (None, "") for value in record.values()):
                    rows.append(record)
            return rows
        finally:
            workbook.close()
