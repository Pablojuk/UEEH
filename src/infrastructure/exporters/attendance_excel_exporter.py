from __future__ import annotations
from pathlib import Path
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class AttendanceExcelExporter:
    def export_attendance_quarterly_excel(self, output_path:str, context:dict[str,Any], rows:list[dict[str,Any]], stats:dict[str,Any])->str:
        path=Path(output_path).expanduser().resolve(); path.parent.mkdir(parents=True,exist_ok=True)
        wb=Workbook(); ws=wb.active; ws.title='Asistencia trimestral'; ws.page_setup.orientation=ws.ORIENTATION_LANDSCAPE
        thin=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        ws.merge_cells('A1:I1'); ws['A1']=context.get('institucion_nombre','Institución'); ws['A1'].font=Font(size=14,bold=True); ws['A1'].alignment=Alignment(horizontal='center')
        ws.merge_cells('A2:I2'); ws['A2']='INFORME TRIMESTRAL DE ASISTENCIA'; ws['A2'].font=Font(size=12,bold=True); ws['A2'].alignment=Alignment(horizontal='center')
        ws['A4']='Asignación'; ws['B4']=f"{context.get('asignatura','')} | {context.get('curso','')}-{context.get('paralelo','')}"
        ws['A5']='Trimestre'; ws['B5']=context.get('trimestre',''); ws['D5']='Período'; ws['E5']=context.get('periodo','')
        ws['A6']='Docente'; ws['B6']=context.get('docente',''); ws['D6']='Año lectivo'; ws['E6']=context.get('periodo_id','')
        headers=['N°','Nómina','Días laborables','Presentes','Atrasos','Faltas injustificadas','Faltas justificadas','% Asistencia','Observación']
        row0=8; ws.append([]); ws.append(headers)
        fill=PatternFill('solid',fgColor='1F4E79'); fnt=Font(color='FFFFFF',bold=True)
        for c in ws[row0]: c.fill=fill; c.font=fnt; c.alignment=Alignment(horizontal='center'); c.border=thin
        for r in rows: ws.append([r['nro'],r['nomina'],r['dias_laborables'],r['presentes'],r['atrasos'],r['faltas_injustificadas'],r['faltas_justificadas'],float(r['porcentaje_asistencia'])/100.0,r['observacion']])
        for i,row in enumerate(ws.iter_rows(min_row=row0+1,max_row=row0+len(rows),min_col=1,max_col=9), start=0):
            for c in row: c.border=thin
            row[7].number_format='0.00%'
            row[8].fill=PatternFill('solid', fgColor=self._obs_fill(rows[i].get('observacion','')))
        r0=row0+len(rows)+2
        ws[f'A{r0}']='Resumen estadístico'; ws[f'A{r0}'].font=Font(bold=True)
        ws[f'A{r0+1}']='Presentes'; ws[f'B{r0+1}']=stats.get('total_presentes',0)
        ws[f'A{r0+2}']='Atrasos'; ws[f'B{r0+2}']=stats.get('total_atrasos',0)
        ws[f'A{r0+3}']='Faltas injustificadas'; ws[f'B{r0+3}']=stats.get('total_faltas_injustificadas',0)
        ws[f'A{r0+4}']='Faltas justificadas'; ws[f'B{r0+4}']=stats.get('total_faltas_justificadas',0)
        ws[f'A{r0+6}']='Porcentaje general'; ws[f'B{r0+6}']=stats.get('porcentaje_general_asistencia',0)/100.0; ws[f'B{r0+6}'].number_format='0.00%'
        ws[f'F{r0+8}']=context.get('firma_docente',''); ws[f'F{r0+9}']='Docente'; ws[f'H{r0+8}']=context.get('firma_rector',''); ws[f'H{r0+9}']='Rector'
        for col,w in [('A',7),('B',36),('C',14),('D',10),('E',10),('F',20),('G',18),('H',12),('I',14)]: ws.column_dimensions[col].width=w
        wb.save(path); return str(path)


    def export_attendance_annual_excel(self, output_path:str, context:dict[str,Any], rows:list[dict[str,Any]], stats:dict[str,Any])->str:
        path=Path(output_path).expanduser().resolve(); path.parent.mkdir(parents=True,exist_ok=True)
        wb=Workbook(); ws=wb.active; ws.title='Asistencia anual'; ws.page_setup.orientation=ws.ORIENTATION_LANDSCAPE
        headers=['N°','Nómina','T1 Días','T1 F','T1 %','T2 Días','T2 F','T2 %','T3 Días','T3 F','T3 %','Días','P','A','F','J','% Asistencia','Observación']
        ws.append(headers)
        for r in rows:
            ws.append([r['nro'],r['nomina'],r['t1_dias'],r['t1_faltas'],r['t1_porcentaje']/100.0,r['t2_dias'],r['t2_faltas'],r['t2_porcentaje']/100.0,r['t3_dias'],r['t3_faltas'],r['t3_porcentaje']/100.0,r['dias_total'],r['presentes_total'],r['atrasos_total'],r['faltas_injustificadas_total'],r['faltas_justificadas_total'],r['porcentaje_asistencia_anual']/100.0,r['observacion']])
        for i in range(len(rows)):
            ws.cell(row=i+2,column=18).fill=PatternFill('solid', fgColor=self._obs_fill(rows[i].get('observacion','')))
            ws.cell(row=i+2,column=17).number_format='0.00%'
        ws['A'+str(len(rows)+3)]='Porcentaje general anual'; ws['B'+str(len(rows)+3)]=stats.get('porcentaje_general_anual_asistencia',0)/100.0; ws['B'+str(len(rows)+3)].number_format='0.00%'
        wb.save(path); return str(path)


    def _obs_fill(self, text: str) -> str:
        t = str(text or '').lower()
        if 'riesgo' in t:
            return 'FCE4D6'
        if 'alerta' in t:
            return 'FED7AA'
        if 'seguimiento' in t:
            return 'FFF2CC'
        return 'E2EFDA'
