"""Exportador de reportes académicos a Excel (.xlsx) con diseño institucional."""

from __future__ import annotations

from pathlib import Path
from typing import Any


class ExcelReportExporter:
    COLOR_HEADER_BG = "1F4E79"
    COLOR_HEADER_FG = "FFFFFF"
    COLOR_FILA_PAR = "DCE6F1"
    COLOR_FILA_IMPAR = "FFFFFF"
    COLOR_APROBADO = "C6EFCE"
    COLOR_REPROBADO = "FFC7CE"
    COLOR_SPL = "FFEB9C"
    COLOR_BORDE = "B8CCE4"

    def exportar(
        self,
        output_path: str,
        report_title: str,
        context: dict[str, Any],
        rows: list[dict[str, Any]],
        ocultar_filas_vacias: bool = False,
    ) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("openpyxl no está instalado") from exc

        filtered_rows = list(rows)
        if ocultar_filas_vacias:
            filtered_rows = [row for row in rows if str(row.get("estudiante", "")).strip()]

        if not filtered_rows and ocultar_filas_vacias:
            filtered_rows = []

        if not rows:
            raise ValueError("No hay datos para exportar")

        path = Path(output_path).expanduser().resolve()
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Informe"

        thin = Side(style="thin", color=self.COLOR_BORDE)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells("A1:N1")
        ws["A1"] = (context.get("institucion_nombre") or "Institución").upper()
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:N2")
        ws["A2"] = report_title.upper()
        ws["A2"].font = Font(bold=True, size=13)
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:N3")
        ws["A3"] = "San Salvador de Cañaribamba"
        ws["A3"].font = Font(bold=False, size=10)
        ws["A3"].alignment = Alignment(horizontal="center")

        self._add_logos(ws, context)

        info_font = Font(bold=True, italic=True)
        ws.merge_cells("A5:B5"); ws["A5"] = "Docente:"; ws["A5"].font = info_font
        ws.merge_cells("C5:D5"); ws["C5"] = context.get("docente_nombre", "N/D")
        ws.merge_cells("A6:B6"); ws["A6"] = "Curso:"; ws["A6"].font = info_font
        ws.merge_cells("C6:D6"); ws["C6"] = context.get("curso_nombre", "N/D")
        ws.merge_cells("A7:B7"); ws["A7"] = "Nivel:"; ws["A7"].font = info_font
        ws.merge_cells("C7:D7"); ws["C7"] = context.get("curso_nivel", "N/D")

        ws.merge_cells("F5:G5"); ws["F5"] = "Asignatura:"; ws["F5"].font = info_font
        ws.merge_cells("H5:I5"); ws["H5"] = context.get("asignatura_nombre", "N/D")
        ws.merge_cells("F6:G6"); ws["F6"] = "Paralelo:"; ws["F6"].font = info_font
        ws.merge_cells("H6:I6"); ws["H6"] = context.get("paralelo_nombre", "N/D")
        ws.merge_cells("F7:G7"); ws["F7"] = "Tipo de reporte:"; ws["F7"].font = info_font
        ws.merge_cells("H7:I7")
        ws["H7"] = "TRIMESTRAL" if context.get("report_type") == "trimestral" else "ANUAL"

        ws.merge_cells("K5:L5"); ws["K5"] = "Tutor:"; ws["K5"].font = info_font
        ws.merge_cells("M5:N5"); ws["M5"] = context.get("firmantes", {}).get("tutor_curso", "")

        for row in range(5, 8):
            for col in [1, 3, 6, 8, 11, 13]:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center")

        start_row = 10
        if context.get("report_type") == "trimestral":
            is_egb = bool(filtered_rows and "promedio_trimestral" in filtered_rows[0])
            if is_egb:
                headers = ["N°", "Nómina", "Promedio Trimestral", "Cualitativo", "Equivalencia", "Observación"]
            else:
                headers = [
                    "N°", "Nómina",
                    "Aportes/Insumos Calificación", "Aportes/Insumos 70%",
                    "Evaluaciones sumativas Calificación", "Evaluaciones sumativas 30%",
                    "Promedio Final", "Cualitativa", "Equivalencia", "Observación",
                ]
            for col, title in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col, value=title)
                cell.font = Font(bold=True, color=self.COLOR_HEADER_FG, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color=self.COLOR_HEADER_BG, end_color=self.COLOR_HEADER_BG, fill_type="solid")
                cell.border = border

            for idx, row in enumerate(filtered_rows, start=1):
                r = start_row + idx
                values = [idx, row.get("estudiante", "")]
                if is_egb:
                    values.extend([row.get("promedio_trimestral"), row.get("cualitativa", ""), row.get("equivalencia", ""), row.get("observacion", "")])
                else:
                    values.extend([
                        row.get("aportes_calificacion"), row.get("aportes_70"), row.get("sumativas_calificacion"), row.get("sumativas_30"),
                        row.get("promedio_final"), row.get("cualitativa", ""), row.get("equivalencia", ""), row.get("observacion", ""),
                    ])
                for cidx, value in enumerate(values, start=1):
                    cell = ws.cell(row=r, column=cidx, value=value)
                    cell.alignment = Alignment(horizontal="left" if cidx == 2 else "center", vertical="center")
                    cell.border = border
                    base_color = self.COLOR_FILA_PAR if idx % 2 == 0 else self.COLOR_FILA_IMPAR
                    cell.fill = PatternFill(start_color=base_color, end_color=base_color, fill_type="solid")
                    if isinstance(value, (int, float)) and cidx != 1:
                        cell.number_format = "0.00"
                obs_text = str(row.get("observacion", "")).strip().upper()
                obs_color = {"APB": self.COLOR_APROBADO, "REP": self.COLOR_REPROBADO, "SPL": self.COLOR_SPL}.get(obs_text)
                if obs_color:
                    obs_cell = ws.cell(row=r, column=(6 if is_egb else 10))
                    obs_cell.fill = PatternFill(start_color=obs_color, end_color=obs_color, fill_type="solid")
            widths = [5, 34, 14, 12, 12, 14] if is_egb else [5, 34, 14, 12, 15, 12, 10, 10, 10, 12]
            signatures_row = max(start_row + len(filtered_rows) + 3, 39)
        else:
            is_egb = bool(filtered_rows and filtered_rows[0].get("supletorio") is None and "equivalencia" in filtered_rows[0])
            if is_egb:
                headers = [
                    "N°", "Nómina", "T1 Calificación", "T1 Cualitativa", "T2 Calificación", "T2 Cualitativa",
                    "T3 Calificación", "T3 Cualitativa", "Promedio", "Cualitativa", "Equivalencia", "Observación",
                ]
            else:
                headers = [
                    "N°", "Nómina",
                    "Trimestre Cali", "Trimestre Cuali",
                    "Trimestre Cali", "Trimestre Cuali",
                    "Trimestre Cali", "Trimestre Cuali",
                    "Promedio", "Cualitativa", "Supletorio", "Promedio Final", "Cualitativo", "Observación",
                ]
            for col, title in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col, value=title)
                cell.font = Font(bold=True, color=self.COLOR_HEADER_FG, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color=self.COLOR_HEADER_BG, end_color=self.COLOR_HEADER_BG, fill_type="solid")
                cell.border = border
            for idx, row in enumerate(filtered_rows, start=1):
                r = start_row + idx
                values = [idx, row.get("estudiante", ""), row.get("trimestre_1"), row.get("equivalencia_t1", ""), row.get("trimestre_2"), row.get("equivalencia_t2", ""), row.get("trimestre_3"), row.get("equivalencia_t3", ""), row.get("promedio"), row.get("cualitativa_anual", "")]
                if is_egb:
                    values.extend([row.get("equivalencia", ""), row.get("observacion", "")])
                else:
                    values.extend([row.get("supletorio"), row.get("promedio_final"), row.get("cualitativo_final", ""), row.get("observacion", "")])
                for cidx, value in enumerate(values, start=1):
                    cell = ws.cell(row=r, column=cidx, value=value)
                    cell.alignment = Alignment(horizontal="left" if cidx == 2 else "center", vertical="center")
                    cell.border = border
                    base_color = self.COLOR_FILA_PAR if idx % 2 == 0 else self.COLOR_FILA_IMPAR
                    cell.fill = PatternFill(start_color=base_color, end_color=base_color, fill_type="solid")
                    if isinstance(value, (int, float)) and cidx != 1:
                        cell.number_format = "0.00"
                obs_text = str(row.get("observacion", "")).strip().upper()
                obs_color = {"APB": self.COLOR_APROBADO, "REP": self.COLOR_REPROBADO, "SPL": self.COLOR_SPL}.get(obs_text)
                if obs_color:
                    obs_cell = ws.cell(row=r, column=(12 if is_egb else 14))
                    obs_cell.fill = PatternFill(start_color=obs_color, end_color=obs_color, fill_type="solid")
            widths = [5, 34, 11, 11, 11, 11, 11, 11, 10, 10, 16, 12] if is_egb else [5, 34, 11, 11, 11, 11, 11, 11, 10, 10, 10, 11, 11, 12]
            last_student_row = start_row + len(filtered_rows)
            stats_start = max(last_student_row + 4, 39)
            signatures_row = self._draw_annual_statistics(ws, stats_start, border, filtered_rows, last_student_row)

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 18

        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(ord("A") + idx - 1)].width = w

        signatures_row = signatures_row + 2
        self._apply_print_settings(ws, len(widths), signatures_row)
        self._draw_signatures(ws, signatures_row, len(widths), context.get("firmantes", {}))
        wb.save(str(path))
        return str(path)

    @staticmethod
    def _apply_print_settings(ws, max_cols: int, max_rows: int) -> None:
        from openpyxl.worksheet.page import PageMargins

        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.2, footer=0.2)
        ws.print_area = f"A1:{chr(ord('A') + max_cols - 1)}{max_rows + 3}"

    def _add_logos(self, ws, context: dict[str, Any]) -> None:
        try:
            from openpyxl.drawing.image import Image
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.utils.units import cm_to_EMU
        except Exception:  # noqa: BLE001
            return

        logo_specs = [
            {
                "path": context.get("logo_path"),
                "base_col": 0,
                "base_row": 0,
                "offset_x_cm": 2.0,
                "offset_y_cm": 1.058,
                "width_cm": 2.07,
                "height_cm": 1.95,
            },
            {
                "path": context.get("logo_ministerio_path"),
                "base_col": 11,
                "base_row": 0,
                "offset_x_cm": 0.0,
                "offset_y_cm": 1.058,
                "width_cm": 2.85,
                "height_cm": 1.31,
            },
        ]

        for spec in logo_specs:
            logo_path = spec["path"]
            if not logo_path:
                continue
            path = Path(str(logo_path)).expanduser().resolve()
            if not path.exists():
                continue
            try:
                image = Image(str(path))
                width_emu = cm_to_EMU(spec["width_cm"])
                height_emu = cm_to_EMU(spec["height_cm"])
                marker = AnchorMarker(
                    col=spec["base_col"],
                    row=spec["base_row"],
                    colOff=cm_to_EMU(spec["offset_x_cm"]),
                    rowOff=cm_to_EMU(spec["offset_y_cm"]),
                )
                image.anchor = OneCellAnchor(_from=marker, ext=(width_emu, height_emu))
                ws.add_image(image)
            except Exception:  # noqa: BLE001
                continue

    def _draw_annual_statistics(self, ws, start_row: int, border, rows: list[dict[str, Any]], last_student_row: int) -> int:
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.styles import Alignment, Font

        aprobados, reprobados = self._count_observations(rows)
        total = aprobados + reprobados
        pct_aprobados = round((aprobados / total) * 100, 2) if total else 0
        pct_reprobados = round((reprobados / total) * 100, 2) if total else 0
        promedio = self._average_promedio_final(rows)

        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        title_cell = ws.cell(row=start_row, column=1, value="Cuadro de logros en la evaluación de los aprendizajes")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        for col in range(1, 7):
            ws.cell(row=start_row, column=col).border = border

        labels = ["Total aprobados", "Total reprobados"]
        values = [aprobados, reprobados]
        percentages = [pct_aprobados, pct_reprobados]

        for idx, label in enumerate(labels, start=1):
            r = start_row + idx
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.cell(row=r, column=1, value=label).alignment = Alignment(horizontal="left")
            ws.cell(row=r, column=5, value=values[idx - 1]).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=6, value=percentages[idx - 1] / 100).number_format = "0%"
            ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
            for col in range(1, 7):
                ws.cell(row=r, column=col).border = border

        promedio_row = start_row + 4
        ws.merge_cells(start_row=promedio_row, start_column=1, end_row=promedio_row, end_column=5)
        ws.cell(row=promedio_row, column=1, value="Promedio").alignment = Alignment(horizontal="center")
        ws.cell(row=promedio_row, column=6, value=promedio if promedio is not None else "—")
        if promedio is not None:
            ws.cell(row=promedio_row, column=6).number_format = "0.00"
        ws.cell(row=promedio_row, column=6).alignment = Alignment(horizontal="center")
        for col in range(1, 7):
            ws.cell(row=promedio_row, column=col).border = border

        chart = PieChart()
        data = Reference(ws, min_col=5, min_row=start_row + 1, max_row=start_row + 2)
        cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + 2)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.height = 3.04 / 2.54
        chart.width = 7.16 / 2.54
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart_row = last_student_row + 2
        ws.add_chart(chart, f"H{chart_row}")

        return max(promedio_row + 3, chart_row + 6)

    @staticmethod
    def _count_observations(rows: list[dict[str, Any]]) -> tuple[int, int]:
        aprobados = 0
        reprobados = 0
        for row in rows:
            obs = str(row.get("observacion", "")).strip().lower()
            has_final = row.get("promedio_final") is not None
            if obs in {"aprobado", "apr", "apb"}:
                aprobados += 1
            elif obs or has_final:
                reprobados += 1
        return aprobados, reprobados

    @staticmethod
    def _average_promedio_final(rows: list[dict[str, Any]]) -> float | None:
        values: list[float] = []
        for row in rows:
            value = row.get("promedio_final")
            if value is None:
                continue
            try:
                values.append(float(value))
            except Exception:  # noqa: BLE001
                continue
        if not values:
            return None
        return round(sum(values) / len(values), 2)

    @staticmethod
    def _draw_signatures(ws, start_row: int, max_cols: int, firmantes: dict[str, str]) -> None:
        from openpyxl.styles import Alignment

        roles = [
            ("Docente", firmantes.get("docente", "")),
            ("Coordinador de Área", firmantes.get("coordinador_area", "")),
            ("Rector", firmantes.get("rector", "")),
            ("Tutor de Curso", firmantes.get("tutor_curso", "")),
        ]
        block = max(1, max_cols // 4)
        for idx, (rol, firma) in enumerate(roles):
            col_start = (idx * block) + 1
            col_end = min(max_cols, col_start + block - 1)
            ws.merge_cells(start_row=start_row, start_column=col_start, end_row=start_row, end_column=col_end)
            ws.merge_cells(start_row=start_row + 1, start_column=col_start, end_row=start_row + 1, end_column=col_end)
            ws.merge_cells(start_row=start_row + 2, start_column=col_start, end_row=start_row + 2, end_column=col_end)
            ws.cell(row=start_row, column=col_start, value="_____________________________")
            ws.cell(row=start_row + 1, column=col_start, value=firma or "")
            ws.cell(row=start_row + 2, column=col_start, value=rol)
            ws.cell(row=start_row, column=col_start).alignment = Alignment(horizontal="center")
            ws.cell(row=start_row + 1, column=col_start).alignment = Alignment(horizontal="center")
            ws.cell(row=start_row + 2, column=col_start).alignment = Alignment(horizontal="center")
