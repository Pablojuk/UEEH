"""Pruebas del exportador Excel."""

from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

from src.infrastructure.exporters.excel_report_exporter import ExcelReportExporter


@unittest.skipIf(importlib.util.find_spec("openpyxl") is None, "openpyxl no instalado")
class TestExcelReportExporter(unittest.TestCase):
    def test_generar_archivo_excel_no_vacio_y_encabezados(self) -> None:
        from openpyxl import load_workbook

        exporter = ExcelReportExporter()
        rows = [
            {
                "estudiante": "Lopez Maria",
                "trimestre_1": 8,
                "trimestre_2": 7,
                "trimestre_3": 6,
                "promedio_final": 7,
                "cualitativo": "B-",
                "observacion": "APB",
                "supletorio": None,
                "nota_definitiva": 7,
            }
        ]
        context = {"contexto_display": "AS1", "institucion_nombre": "UEEH", "report_type": "anual"}

        with tempfile.TemporaryDirectory() as tmp:
            output = str(Path(tmp) / "reporte.xlsx")
            result = exporter.exportar(output, "Reporte", context, rows)
            self.assertTrue(Path(result).exists())
            self.assertGreater(Path(result).stat().st_size, 0)

            wb = load_workbook(result)
            ws = wb.active
            self.assertEqual(ws["A6"].value, "N°")
            self.assertEqual(ws["B6"].value, "Nómina")
            wb.close()

    def test_reporte_anual_agrega_cuadro_estadistico_promedio_y_grafica(self) -> None:
        from openpyxl import load_workbook

        exporter = ExcelReportExporter()
        rows = [
            {"estudiante": "A", "promedio_final": 9, "observacion": "APB"},
            {"estudiante": "B", "promedio_final": 5, "observacion": "REP"},
        ]
        context = {"institucion_nombre": "UEEH", "report_type": "anual"}

        with tempfile.TemporaryDirectory() as tmp:
            output = str(Path(tmp) / "reporte.xlsx")
            result = exporter.exportar(output, "Reporte", context, rows)
            wb = load_workbook(result)
            ws = wb.active

            stats_title = None
            for row_idx in range(1, ws.max_row + 1):
                if ws.cell(row=row_idx, column=1).value == "Cuadro de logros en la evaluación de los aprendizajes":
                    stats_title = row_idx
                    break

            self.assertIsNotNone(stats_title)
            self.assertEqual(ws.cell(row=stats_title + 1, column=1).value, "Total aprobados")
            self.assertEqual(ws.cell(row=stats_title + 1, column=5).value, 1)
            self.assertEqual(ws.cell(row=stats_title + 2, column=1).value, "Total reprobados")
            self.assertEqual(ws.cell(row=stats_title + 2, column=5).value, 1)
            self.assertEqual(ws.cell(row=stats_title + 4, column=1).value, "Promedio")
            self.assertEqual(ws.cell(row=stats_title + 4, column=6).value, 7)
            self.assertGreaterEqual(len(ws._charts), 1)
            wb.close()


if __name__ == "__main__":
    unittest.main()
